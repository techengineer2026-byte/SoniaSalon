from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from datetime import date, timedelta
import copy

wb = Workbook()

STAFF = ['Sonia', 'Sandeep', 'Sigi', 'Sangeta']

# Generate next 14 days
TODAY = date.today()
DATES = [TODAY + timedelta(days=i) for i in range(14)]

# Time slots 9am - 8pm every 30 min
def gen_slots():
    slots = []
    for h in range(9, 20):
        for m in (0, 30):
            suffix = 'AM' if h < 12 else 'PM'
            hh = h if h <= 12 else h - 12
            if h == 12: hh = 12
            slots.append(f"{hh}:{m:02d} {suffix}")
    return slots

TIME_SLOTS = gen_slots()

# ── Colors ──────────────────────────────────────────────────
PINK_DARK   = 'B76E79'
PINK_LIGHT  = 'F5E6E8'
GREEN_BG    = 'E8F5E9'
RED_BG      = 'FFEBEE'
HEADER_BG   = '2C2C2C'
ALT_ROW     = 'FAF7F7'
WHITE       = 'FFFFFF'
GRAY_BORDER = 'D0D0D0'

thin = Side(style='thin', color=GRAY_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(bold=True, color=WHITE, size=11):
    return Font(name='Arial', bold=bold, color=color, size=size)

def cell_font(bold=False, color='2C2C2C', size=10):
    return Font(name='Arial', bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def make_staff_sheet(staff_name):
    ws = wb.create_sheet(title=staff_name)

    # ── Title row ──────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(DATES)+1)
    title_cell = ws.cell(row=1, column=1,
                         value=f"{staff_name} — Weekly Schedule")
    title_cell.font    = Font(name='Arial', bold=True, color=WHITE, size=13)
    title_cell.fill    = fill(PINK_DARK)
    title_cell.alignment = center()
    ws.row_dimensions[1].height = 28

    # ── Sub-header: instructions ───────────────────────────
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=len(DATES)+1)
    inst = ws.cell(row=2, column=1,
                   value="✓ = Available (client can book)   ✗ = Blocked / day off   "
                         "Check the box to make a slot available.")
    inst.font      = Font(name='Arial', italic=True, color='666666', size=9)
    inst.fill      = fill('FDF0F1')
    inst.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # ── Column headers: Time | Date1 | Date2 … ────────────
    days = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
    months = ['Jan','Feb','Mar','Apr','May','Jun',
              'Jul','Aug','Sep','Oct','Nov','Dec']

    # Time column header
    tc = ws.cell(row=3, column=1, value='Time')
    tc.font = hdr_font()
    tc.fill = fill(HEADER_BG)
    tc.alignment = center()
    tc.border = border
    ws.column_dimensions['A'].width = 11

    for col_idx, d in enumerate(DATES, start=2):
        label = f"{days[d.weekday()]}\n{d.day} {months[d.month-1]}"
        hc = ws.cell(row=3, column=col_idx, value=label)
        hc.font = hdr_font(size=10)
        # Weekend = slightly different shade
        bg = 'E8C5CB' if d.weekday() >= 5 else PINK_DARK
        hc.fill = fill(bg)
        hc.alignment = center()
        hc.border = border
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 10

    ws.row_dimensions[3].height = 34

    # ── Time slot rows with checkboxes ────────────────────
    for row_idx, slot in enumerate(TIME_SLOTS, start=4):
        # Alternate row background
        row_bg = WHITE if row_idx % 2 == 0 else ALT_ROW

        # Time label
        tc = ws.cell(row=row_idx, column=1, value=slot)
        tc.font      = cell_font(bold=True, color='555555')
        tc.fill      = fill('F0E8EA')
        tc.alignment = center()
        tc.border    = border

        for col_idx in range(2, len(DATES)+2):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Default all to TRUE (available) — Sonia unchecks to block
            cell.value     = True
            cell.data_type = 'b'
            cell.fill      = fill(row_bg)
            cell.alignment = center()
            cell.border    = border

        ws.row_dimensions[row_idx].height = 18

    # ── Conditional formatting: green if TRUE, red if FALSE ─
    last_col = get_column_letter(len(DATES)+1)
    data_range = f"B4:{last_col}{3+len(TIME_SLOTS)}"

    ws.conditional_formatting.add(data_range,
        FormulaRule(formula=['B4=TRUE'],
                    fill=PatternFill('solid', start_color=GREEN_BG, fgColor=GREEN_BG),
                    font=Font(color='2E7D32', bold=True)))

    ws.conditional_formatting.add(data_range,
        FormulaRule(formula=['B4=FALSE'],
                    fill=PatternFill('solid', start_color=RED_BG, fgColor=RED_BG),
                    font=Font(color='C62828', bold=True)))

    # ── Legend below data ──────────────────────────────────
    legend_row = 4 + len(TIME_SLOTS) + 1
    ws.merge_cells(start_row=legend_row, start_column=1,
                   end_row=legend_row, end_column=4)
    lc = ws.cell(row=legend_row, column=1,
                 value="Legend:   TRUE / checked = client CAN book   |   "
                       "FALSE / unchecked = BLOCKED")
    lc.font = Font(name='Arial', italic=True, color='888888', size=9)
    lc.alignment = Alignment(horizontal='left', vertical='center')

    # Freeze pane so time column & header stay visible
    ws.freeze_panes = 'B4'

    return ws


# ── Build all staff sheets ─────────────────────────────────
for name in STAFF:
    make_staff_sheet(name)

# Remove default sheet
del wb['Sheet']

# ── HOW TO USE sheet ──────────────────────────────────────
ws_info = wb.create_sheet(title='HOW TO USE', index=0)
ws_info.sheet_properties.tabColor = PINK_DARK

instructions = [
    ("SONIA'S SALON — Staff Schedule Guide", None, True, 14, WHITE, PINK_DARK),
    ("", None, False, 10, '333333', WHITE),
    ("HOW IT WORKS", None, True, 11, WHITE, '555555'),
    ("Each staff member has their own tab (Sonia, Sandeep, Sigi, Sangeta).", None, False, 10, '333333', WHITE),
    ("Each tab shows 14 days across the top and time slots (9am–8pm) down the side.", None, False, 10, '333333', WHITE),
    ("", None, False, 10, '333333', WHITE),
    ("TO BLOCK A SLOT (make it unavailable):", None, True, 10, WHITE, '8B3A3A'),
    ("  Click the checkbox in the cell → it turns RED → clients cannot book that slot.", None, False, 10, '333333', WHITE),
    ("", None, False, 10, '333333', WHITE),
    ("TO OPEN A SLOT (make it available):", None, True, 10, WHITE, '2E7D32'),
    ("  Click the checkbox → it turns GREEN → clients CAN book that slot.", None, False, 10, '333333', WHITE),
    ("", None, False, 10, '333333', WHITE),
    ("BLOCKING A FULL DAY OFF:", None, True, 11, WHITE, '555555'),
    ("  Select all cells in that date column → press Delete → all become FALSE (red).", None, False, 10, '333333', WHITE),
    ("", None, False, 10, '333333', WHITE),
    ("IMPORTANT NOTES", None, True, 11, WHITE, '555555'),
    ("  • Even if a slot is GREEN, a confirmed booking will also block it in Google Calendar.", None, False, 10, '333333', WHITE),
    ("  • The booking system checks BOTH this sheet AND Google Calendar before showing a slot.", None, False, 10, '333333', WHITE),
    ("  • Update this sheet at the start of each week or whenever hours change.", None, False, 10, '333333', WHITE),
    ("  • Do NOT change the column/row structure — the booking system reads exact positions.", None, False, 10, '333333', WHITE),
    ("", None, False, 10, '333333', WHITE),
    ("SHEET NAME = Staff Name (must match exactly what is in the Staff sheet in Google Sheets)", None, True, 10, 'B76E79', WHITE),
]

ws_info.column_dimensions['A'].width = 80
for r, (text, _, bold, size, color, bg) in enumerate(instructions, start=1):
    c = ws_info.cell(row=r, column=1, value=text)
    c.font = Font(name='Arial', bold=bold, color=color, size=size)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal='left', vertical='center',
                            wrap_text=True, indent=1)
    ws_info.row_dimensions[r].height = 20 if text else 8

# ── Save ──────────────────────────────────────────────────
# ── Save ──────────────────────────────────────────────────
out = 'Sonias_Staff_Schedule.xlsx'
wb.save(out)

print("Saved:", out)
